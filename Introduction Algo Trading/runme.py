from funcs.Extractor import extract_stock_price
from funcs.Bollinger import bollinger
from funcs.RSI import rsi
from funcs.Strategy import strategy
from funcs.Graph import plot_startegy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

pd.options.plotting.backend = "plotly"

df = extract_stock_price(ticker='SQ',
                         start='2019-01-01',
                         end='2022-08-13')

bb = bollinger(df_prices=df,
               window=14,
               delta=1.3)

rsi = rsi(df_prices=bb,
          window=14)

result_strat_1 = strategy(df_prices=rsi,
                          rsi_lower=40,
                          rsi_upper=60,
                          stop_loss=False)

result_strat_2 = strategy(df_prices=rsi,
                          rsi_lower=40,
                          rsi_upper=60,
                          stop_loss=True,
                          sl=0.70)

fig1_strat_1, fig2_strat_1, fig3_strat_1 = plot_startegy(result_strat_1)
fig1_strat_2, fig2_strat_2, fig3_strat_2 = plot_startegy(result_strat_2)

fig3_strat_1.write_image(file='plot_strat_1.png', format='png')
fig3_strat_2.write_image(file='plot_strat_2.png', format='png')

result_strat_1["df"].to_excel('Result of Strategy one.xlsx', sheet_name='Dataset')
result_strat_2["df"].to_excel('Result of Strategy two.xlsx', sheet_name='Dataset')

with pd.ExcelWriter(
        "Result of Strategy one.xlsx",
        mode="a",
        engine="openpyxl",
        if_sheet_exists="replace",
) as writer:
    result_strat_1["buy"].to_excel(writer, sheet_name='buy')
    result_strat_1["sell"].to_excel(writer, sheet_name='sell')

with pd.ExcelWriter(
        "Result of Strategy two.xlsx",
        mode="a",
        engine="openpyxl",
        if_sheet_exists="replace",
) as writer:
    result_strat_2["buy"].to_excel(writer, sheet_name='buy')
    result_strat_2["sell"].to_excel(writer, sheet_name='sell')

wb = load_workbook(f'Result of Strategy one.xlsx')
sheet1 = wb.create_sheet('graph', 0)
active = wb['graph']
active.add_image(Image('plot_strat_1.png'), 'A1')
rtn_sheet = wb.create_sheet('return', 0)
rtn_sheet['A1'] = 'Return of the Strategy'
rtn_sheet['A2'] = result_strat_1['return']

wb.save('Result of Strategy one.xlsx')

wb = load_workbook(f'Result of Strategy two.xlsx')
sheet1 = wb.create_sheet('graph', 0)
active = wb['graph']
active.add_image(Image('plot_strat_2.png'), 'A1')
rtn_sheet = wb.create_sheet('return', 0)
rtn_sheet['A1'] = 'Return of the Strategy'
rtn_sheet['A2'] = result_strat_2['return']
wb.save('Result of Strategy two.xlsx')



